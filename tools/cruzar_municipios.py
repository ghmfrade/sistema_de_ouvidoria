"""
Cruza os municípios das paradas (metropolitano e intermunicipal)
com a tabela municipios (IBGE/SP).

Estratégias de normalização (sem busca por proximidade):
  1. Exato (case-insensitive)
  2. Tudo maiúsculas, sem acentos, sem hífens/apóstrofos
  3. Sem espaços (além da etapa 2)

Saída: cruzamento_municipios.xlsx  (duas abas)
"""

import os
import re
import sys
import unicodedata

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.connection import get_session
from models import Municipio, ParadaAutoLinha
from models.auto_linha import AutoLinha, TipoServico
from sqlalchemy import select


# ── helpers ──────────────────────────────────────────────────────────────────

def normalizar(s: str) -> str:
    """Maiúsculas + sem acentos + sem hífen/apóstrofo."""
    s = s.upper().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")  # remove combining marks
    s = re.sub(r"[-'`]", " ", s)          # hífen/apóstrofo → espaço
    s = re.sub(r"\s+", " ", s).strip()    # normaliza espaços
    return s


def normalizar_sem_espaco(s: str) -> str:
    """Igual a normalizar() mas sem nenhum espaço."""
    return normalizar(s).replace(" ", "")


def match_estrategia(nome_parada: str, ibge_exact: set, ibge_norm: dict, ibge_noesp: dict):
    """
    Retorna (cod_ibge, nome_ibge, estrategia) ou (None, None, 'SEM MATCH').

    ibge_exact  : {nome_lower -> (cod_ibge, nome_ibge)}
    ibge_norm   : {normalizado -> (cod_ibge, nome_ibge)}
    ibge_noesp  : {sem_espaco -> (cod_ibge, nome_ibge)}
    """
    # Estratégia 1: exato (case-insensitive)
    k1 = nome_parada.strip().lower()
    if k1 in ibge_exact:
        cod, nome = ibge_exact[k1]
        return cod, nome, "1-exato"

    # Estratégia 2: normalizado (sem acentos, sem hífen)
    k2 = normalizar(nome_parada)
    if k2 in ibge_norm:
        cod, nome = ibge_norm[k2]
        return cod, nome, "2-normalizado"

    # Estratégia 3: sem espaço
    k3 = normalizar_sem_espaco(nome_parada)
    if k3 in ibge_noesp:
        cod, nome = ibge_noesp[k3]
        return cod, nome, "3-sem-espaco"

    return None, None, "SEM MATCH"


# ── busca no banco ────────────────────────────────────────────────────────────

session = get_session()

# Municípios IBGE (SP)
municipios = session.execute(
    select(Municipio.cod_ibge, Municipio.nome)
    .where(Municipio.estado == "SP")
    .order_by(Municipio.nome)
).all()

ibge_exact = {m.nome.strip().lower(): (m.cod_ibge, m.nome) for m in municipios}
ibge_norm  = {normalizar(m.nome): (m.cod_ibge, m.nome) for m in municipios}
ibge_noesp = {normalizar_sem_espaco(m.nome): (m.cod_ibge, m.nome) for m in municipios}

print(f"Municípios IBGE/SP carregados: {len(municipios)}")

# Cidades metropolitanas (distinct)
cidades_metro = session.execute(
    select(ParadaAutoLinha.cidade)
    .join(AutoLinha, AutoLinha.id == ParadaAutoLinha.auto_id)
    .where(AutoLinha.tipo == TipoServico.REGULAR_METROPOLITANO)
    .distinct()
    .order_by(ParadaAutoLinha.cidade)
).scalars().all()

print(f"Cidades metropolitanas (distinct): {len(cidades_metro)}")

# Cidades intermunicipais (distinct)
cidades_inter = session.execute(
    select(ParadaAutoLinha.cidade)
    .join(AutoLinha, AutoLinha.id == ParadaAutoLinha.auto_id)
    .where(AutoLinha.tipo == TipoServico.REGULAR_INTERMUNICIPAL)
    .distinct()
    .order_by(ParadaAutoLinha.cidade)
).scalars().all()

print(f"Cidades intermunicipais (distinct): {len(cidades_inter)}")

session.close()

# ── cruzamento ────────────────────────────────────────────────────────────────

def cruzar(cidades: list[str]) -> pd.DataFrame:
    rows = []
    for cidade in cidades:
        cod, nome_ibge, estrategia = match_estrategia(cidade, ibge_exact, ibge_norm, ibge_noesp)
        rows.append({
            "nome_parada":   cidade,
            "cod_ibge":      cod,
            "nome_ibge":     nome_ibge,
            "estrategia":    estrategia,
        })
    df = pd.DataFrame(rows)
    # Ordenar: sem match primeiro para facilitar revisão
    ordem = {"SEM MATCH": 0, "1-exato": 1, "2-normalizado": 2, "3-sem-espaco": 3}
    df["_ord"] = df["estrategia"].map(ordem)
    df = df.sort_values(["_ord", "nome_parada"]).drop(columns="_ord").reset_index(drop=True)
    return df


df_metro = cruzar(cidades_metro)
df_inter = cruzar(cidades_inter)

# ── resumo ────────────────────────────────────────────────────────────────────

def resumo(df: pd.DataFrame, label: str):
    total = len(df)
    sem = (df["estrategia"] == "SEM MATCH").sum()
    print(f"\n{label}")
    print(f"  Total cidades: {total}")
    print(f"  Com match:     {total - sem} ({100*(total-sem)/total:.1f}%)")
    print(f"  Sem match:     {sem} ({100*sem/total:.1f}%)")
    for e in ["1-exato", "2-normalizado", "3-sem-espaco"]:
        n = (df["estrategia"] == e).sum()
        print(f"    {e}: {n}")
    if sem > 0:
        print("  >>> Sem match:")
        for nome in df[df["estrategia"] == "SEM MATCH"]["nome_parada"]:
            print(f"       {nome!r}")

resumo(df_metro, "METROPOLITANO")
resumo(df_inter, "INTERMUNICIPAL")

# ── exportar Excel ────────────────────────────────────────────────────────────

out_path = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "cruzamento_municipios.xlsx"
)

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df_metro.to_excel(writer, sheet_name="Metropolitano", index=False)
    df_inter.to_excel(writer, sheet_name="Intermunicipal", index=False)

    # Formatação básica — ajusta largura das colunas
    for sheet_name, df in [("Metropolitano", df_metro), ("Intermunicipal", df_inter)]:
        ws = writer.sheets[sheet_name]
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        # Colorir linhas sem match de vermelho claro
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for row in ws.iter_rows(min_row=2):
            if row[3].value == "SEM MATCH":   # coluna estrategia
                for cell in row:
                    cell.fill = red_fill

print(f"\nArquivo gerado: {out_path}")
